import tkinter as tk
from tkinter import ttk
import openpyxl
import os
import time
from tkinter import messagebox

'''
This program creates a GUI to enter student data into an Excel file.
The program uses the openpyxl module to read and write Excel files.
The program uses the tkinter module to create the GUI.
Up to now the program:
- Can not create a new data.xlsx file. We must create it manually.
- Must be run from the same directory as the data.xlsx file.
- Does not check if the data.xlsx file is open by another user.
- Does not check if the data.xlsx file exists.
- Does not check if the data.xlsx file is empty and does not have column headers. 
- It should write the column headers if they are missing.
- Does not check if the data.xlsx file has the correct column headers.\
- Does not check if the data.xlsx file has the correct number of columns.
- The menu options are given by the schools.xlsx file. The schools.xlsx file must be in the same directory as the script.
- The schools.xlsx file must have a sheet named "Schools".
- The schools.xlsx file must have a column named "School Name".
- The schools.xlsx file must have a column named "Class 1", "Class 2", "Class 3", etc.
- The schools.xlsx file must have a row for each school.
- The schools.xlsx file must have a row for each class.
- When the school is changed, the class dropdown menu is not updated.
- The program does not check if the school and class are valid.
- The program does not check if the school and class are selected.
- The program does not check if the name and grade are entered.
- The program does not check if the name and grade are valid.

- The program should show the newly added data in the data viewer.
- The program should allow the user to double-click a row in the viewer and the row data to go for editing in the entry form
- The program should allow the user to edit the data in the entry form and save the changes to the Excel file.
- The program should allow the user to delete a row from the Excel file.
- The program should allow the user to sort the data in the viewer by any column.
- The program should allow the user to filter the data in the viewer by any column.
- The program should allow the user to search the data in the viewer by any column.
- The program should allow the user to export the data in the viewer to a CSV file.
- The program should allow the user to import data from a CSV file to the Excel file.
- The program should allow the user to create a new Excel file.
- The program should allow the user to create a new Excel file from a CSV file.




'''


class StudentEntry:
    def __init__(self, parent):
        self.parent = parent
        self.create_widgets(parent)
        self.create_data_file_if_not_exists()
        self.create_menu()
        self.student_entry = None  # Reference to the StudentEntry object

    def create_data_file_if_not_exists(self):
        if not os.path.isfile('data.xlsx'):
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            headers = ['Name', 'Grade', 'Gender', 'OldSchool', 'OldSchoolClass']
            sheet.append(headers)

            workbook.save('data.xlsx')

    def create_widgets(self,parent):
    # Create the labels
        label_name = tk.Label(parent, text="Name:")
        label_grade = tk.Label(parent, text="Grade:")
        label_gender = tk.Label(parent, text="Gender:")
        label_old_school = tk.Label(parent, text="Old School:")
        label_old_school_class = tk.Label(parent, text="Old School Class:")

        # Continue with the other steps in creating the entry form
    



        # Position the labels using a layout manager
        label_name.grid(row=0, column=0, sticky=tk.W)
        label_grade.grid(row=1, column=0, sticky=tk.W)
        label_gender.grid(row=2, column=0, sticky=tk.W)
        label_old_school.grid(row=3, column=0, sticky=tk.W)
        label_old_school_class.grid(row=4, column=0, sticky=tk.W)

        # Create Entry Widgets
        self.name_entry = tk.Entry(parent)  # Entry widget for name

        # Create the Submit Button
        submit_button = tk.Button(parent, text="Submit", command=self.submit)

        # Add drop-down menus for Grade, Gender, OldSchool, and nested menu for OldSchoolClass
        # Create the drop-down menus
        grade_var = tk.StringVar()
        gender_var = tk.StringVar()
        old_school_var = tk.StringVar()
        old_school_class_var = tk.StringVar()

        grade_combobox = ttk.Combobox(parent, textvariable=grade_var, values=['A', 'B', 'C'])
        gender_combobox = ttk.Combobox(parent, textvariable=gender_var, values=['Male', 'Female'])
        old_school_combobox = ttk.Combobox(parent, textvariable=old_school_var)
        old_school_class_combobox = ttk.Combobox(parent, textvariable=old_school_class_var)

        # Set the options for OldSchool by calling get_school_menu_options method
        old_school_options = self.get_school_menu_options()
        old_school_combobox['values'] = old_school_options

        # Nest the OldSchoolClass menu based on OldSchool selection
        def update_school_class_options(event):
            selected_school = old_school_var.get()
            # Retrieve the corresponding OldSchoolClass options based on selected_school
            # Set the options for OldSchoolClass combobox
            old_school_class_options = self.get_school_class_options(selected_school)
            old_school_class_combobox['values'] = old_school_class_options

        old_school_combobox.bind('<<ComboboxSelected>>', update_school_class_options)

        # Position Labels and Entry Widgets on the grid
        label_name.grid(row=0, column=0, sticky="e")
        self.name_entry.grid(row=0, column=1)

        # Position the drop-down menus using a layout manager
        grade_combobox.grid(row=1, column=1, sticky=tk.W)
        gender_combobox.grid(row=2, column=1, sticky=tk.W)
        old_school_combobox.grid(row=3, column=1, sticky=tk.W)
        old_school_class_combobox.grid(row=4, column=1, sticky=tk.W)
        label_name.grid(row=5, column=0, sticky="e")
        self.name_entry.grid(row=5, column=1)

        submit_button.grid(row=6, column=1)  # Adjust the row and column as per your layout

        # Create the submit and cancel buttons

        # Set the placement of the object in a grid of the parent window

    def get_school_menu_options(self):
        workbook = openpyxl.load_workbook('schools.xlsx')
        sheet = workbook.active

        options = []
        for row in sheet.iter_rows(values_only=True):
            school_name = row[0]
            options.append(school_name)

        return options

    def get_school_class_options(self, selected_school):
        workbook = openpyxl.load_workbook('schools.xlsx')
        sheet = workbook.active

        options = []
        for row in sheet.iter_rows(values_only=True):
            school_name = row[0]
            school_class_options = row[1:]

            if school_name == selected_school:
                options.extend(school_class_options)

        return options
    
    def create_menu(self):
        submit_button = tk.Button(self.parent, text="Submit", command=self.submit)
        submit_button.grid(row=6, column=0, padx=10, pady=10)

        cancel_button = tk.Button(self.parent, text="Cancel", command=self.cancel)
        cancel_button.grid(row=6, column=1, padx=10, pady=10)


    def submit(self):
        name = self.name_entry.get()
        grade = self.grade_combobox.get()
        gender = self.gender_combobox.get()
        school = self.school_combobox.get()
        class_ = self.class_combobox.get()
        while True:
            try:
                self.wb = openpyxl.load_workbook("data.xlsx")
                self.ws = self.wb.active
                self.ws.append([name, grade, gender, school, class_])
                self.wb.save("data.xlsx")
                self.submit_button.config(state='normal')
                self.display_student.display_data()  # Call method to update display
                break
            except PermissionError:
                self.submit_button.config(state='disabled')
                messagebox.showerror("Excel file is open", "Please close the Excel file and try again.")
                time.sleep(0.1)


        # # Save the student record to data.xlsx
        # workbook = openpyxl.load_workbook('data.xlsx')
        # sheet = workbook.active

        row = [name, grade, gender, old_school, old_school_class]
        self.ws.append(row)

        self.ws.save('data.xlsx')

    def cancel(self):
        self.name_entry.delete(0, tk.END)
        self.grade_combobox.set('')
        self.gender_combobox.set('')
        self.old_school_combobox.set('')
        self.old_school_class_combobox.set('')

    def re_edit_student(self, student_record):
        name, grade, gender, old_school, old_school_class = student_record

        self.name_entry.delete(0, tk.END)
        self.name_entry.insert(0, name)

        self.grade_combobox.set(grade)
        self.gender_combobox.set(gender)
        self.old_school_combobox.set(old_school)
        self.old_school_class_combobox.set(old_school_class)

class StudentsDisplay:
    def __init__(self, parent):
        self.parent = parent
        #self.create_widgets()
        self.create_data_file_if_not_exists()
        self.treeview = ttk.Treeview(self.parent)
        self.student_records = []  # Initialize as an empty list
        
        # Load and display student records
        self.load_student_records()
        self.display_student_records(self.student_records)

    def on_row_double_click(self, event):
        selected_item = self.treeview.focus()  # Get the selected item
        record_values = self.treeview.item(selected_item, 'values')
        # Start the editing process using the record_values
        if self.student_entry:
            self.student_entry.re_edit_student(record_values)

    def create_data_file_if_not_exists(self):
        if not os.path.isfile('data.xlsx'):
            workbook = openpyxl.Workbook()
            workbook.save('data.xlsx')    

   # Method to load and display student records from data.xlsx
    def load_student_records(self):
        # Load data from data.xlsx
        workbook = openpyxl.load_workbook('data.xlsx')
        sheet = workbook.active

        # Retrieve student records from the sheet
        student_records = []
        for row in sheet.iter_rows(values_only=True):
            student_records.append(row)

        # Display student records in the rows
        self.display_student_records(student_records)

    # Method to display student records in the rows
    def display_student_records(self, student_records):
        # Clear existing rows (if any)
        self.treeview.delete(*self.treeview.get_children())

        # Configure columns and headers
        self.treeview['columns'] = ('Name', 'Grade', 'Gender', 'OldSchool', 'OldSchoolClass')
        self.treeview.heading('#0', text='ID')
        self.treeview.heading('Name', text='Name')
        self.treeview.heading('Grade', text='Grade')
        self.treeview.heading('Gender', text='Gender')
        self.treeview.heading('OldSchool', text='Old School')
        self.treeview.heading('OldSchoolClass', text='Old School Class')

        # Display student records in rows
        for index, record in enumerate(student_records):
            self.treeview.insert('', 'end', text=index+1, values=record)

        # Other Treeview formatting and configuration 
        # Set column widths
        self.treeview.column('#0', width=50)
        self.treeview.column('Name', width=150)
        self.treeview.column('Grade', width=80)
        self.treeview.column('Gender', width=80)
        self.treeview.column('OldSchool', width=120)
        self.treeview.column('OldSchoolClass', width=120)


        # Create vertical scrollbar
        scrollbar = ttk.Scrollbar(self.parent, orient='vertical')
        scrollbar.configure(command=self.treeview.yview)
        self.treeview.configure(yscrollcommand=scrollbar.set)

  
        # Place the Treeview and scrollbar on the grid
        self.treeview.grid(row=0, column=0, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')

        # Configure grid weights to allow resizing of the Treeview
        self.parent.grid_rowconfigure(0, weight=1)
        self.parent.grid_columnconfigure(0, weight=1)

        # MORE Treeview formatting and configuration
#
        # Bind double-click event to the Treeview
        self.treeview.bind('<<TreeviewSelect>>', self.on_row_double_click)



# Main program
if __name__ == "__main__":
    # Create the main window
    root = tk.Tk()
    root.title("Student Records")

    # Create the StudentEntry form and StudentsDisplay widget
    student_entry = StudentEntry(root)
    students_display = StudentsDisplay(root)
    students_display.student_entry = student_entry  # Pass the reference


    # Set the placement of the StudentEntry form and StudentsDisplay widget in the main window

    # Start the Tkinter event loop
    root.mainloop()
