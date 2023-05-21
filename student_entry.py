import tkinter as tk
import openpyxl
import tkinter as tk
import pandas as pd
from tkinter import messagebox
import time


'''
This object creates a GUI to enter student data into an Excel file called data.xlsx or to re_edit the data in the Excel file.
Whenever a submit is made, the data are written in the Excel file and the data viewer of Display_students object is refreshed.
Display_students object can fill the StudentEntry form with the data of an existing student, selected with double-click from the data viewer.
In addition, it changes re_edit status of this object to True, in order the submit method to know that the data must be replaced 
in the Excel file. To do so, the position of the row in the Excel file is injected in this object by the Display_students object.
So, this object has to be injected within the Display_student object by the main program called write_student.py. 
A clear button is added to clear the form fields.
Do not submit if the name, grade, gender, school and class are not selected or entered.
A delete button is added to delete the selected row from the Excel file.

'''

class StudentEntry(tk.Frame):
    def __init__(self, master=None, file_path="data.xlsx"):
        super().__init__(master)
        self.master = master
        self.file_path = file_path
        self.ws = openpyxl.load_workbook(file_path).active
        self.grid()
        self.create_widgets()
        self.re_edit = False


    def create_widgets(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.append(['Name', 'Grade', 'Gender', 'OldSchool', 'OldSchoolClass'])


        # Create the form labels and entry fields
        self.name_label = tk.Label(self, text="Name:")
        self.name_label.grid(row=0, column=0)
        self.name_entry = tk.Entry(self)
        self.name_entry.grid(row=0, column=1)

        self.grade_label = tk.Label(self, text="Grade:")
        self.grade_label.grid(row=1, column=0)
        self.grade_entry = tk.Entry(self)
        self.grade_entry.grid(row=1, column=1)

        self.gender_label = tk.Label(self, text="Gender:")
        self.gender_label.grid(row=2, column=0)
        self.gender_options = ["Male", "Female", "Other"]
        self.gender_var = tk.StringVar(self)
        self.gender_var.set(self.gender_options[0])
        self.gender_dropdown = tk.OptionMenu(self, self.gender_var, *self.gender_options)
        self.gender_dropdown.grid(row=2, column=1)

        self.school_label = tk.Label(self, text="OldSchool:")
        self.school_label.grid(row=3, column=0)
        self.school_options = self.get_school_options() # get options from another Excel file
        self.school_var = tk.StringVar(self)
        self.school_var.set(self.school_options[0])
        self.school_dropdown = tk.OptionMenu(self, self.school_var, *self.school_options, command=self.update_class_dropdown)
        self.school_dropdown.grid(row=3, column=1)

        self.update_class_dropdown(self.school_options[0])
        self.class_label = tk.Label(self, text="OldSchoolClass:")
        self.class_label.grid(row=4, column=0)
        self.class_var = tk.StringVar(self)
        self.class_var.set(self.class_options[0])
        self.class_dropdown = tk.OptionMenu(self, self.class_var, "")
        self.class_dropdown.grid(row=4, column=1)

        # Create the submit button
        self.submit_button = tk.Button(self, text="Submit", command=self.submit)
        self.submit_button.grid(row=5, column=1)

        # Create the quit button
        self.quit_button = tk.Button(self, text="Quit", command=self.master.destroy)
        self.quit_button.grid(row=5, column=0)

        # Create the clear button
        self.clear_button = tk.Button(self, text="Clear", command=self.clear)
        self.clear_button.grid(row=5, column=2)

        # Create the delete button
        self.delete_button = tk.Button(self, text="Delete", command=self.delete)
        self.delete_button.grid(row=5, column=3)

    def get_school_options(self):
        # Read the options from another Excel file
        wb = openpyxl.load_workbook("schools.xlsx")
        ws = wb["Schools"]
        school_options = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            school_options.append(row[0])
        return school_options


    def update_class_dropdown(self, selection):
        print(f"Updating class dropdown for school {selection}")
        print(f"Selected school: {selection}")

        # Load the class options for the selected school from the "schools.xlsx" file
        wb = openpyxl.load_workbook('schools.xlsx')
        ws = wb['Schools']
        self.class_options = []
        for row in ws.iter_rows(min_row=2, min_col=1, values_only=True):
            if row[0] == selection:
                self.class_options = [c for c in row[1:] if c is not None]
                break
        print(f"Class options for {selection}: {self.class_options}")
        print(f"Class options: {self.class_options}")

        try:
            # Clear and repopulate the "OldSchoolClass" dropdown menu
            self.class_dropdown['menu'].delete(0, 'end')
            for option in self.class_options:
                self.class_dropdown['menu'].add_command(label=option, command=tk._setit(self.class_var, option))
        except:
            pass


    def replace_row(self, row_index, new_values):
        # Get the row number from the row index
        print("I'm in replace_row with row_index ",row_index)
        # row_number = int(''.join(filter(str.isdigit, row_index)))
        # # Get the row based on the index (assuming 1-based index)
        row = self.ws[row_index]
        print("row ",row)
        # Replace the values in the row with the new values
        for cell, new_value in zip(row, new_values):
            print("cell new value ",cell, new_value)
            cell.value = new_value

        # Save the workbook
        save_path = "data.xlsx"
        self.wb.save(save_path) 
   

    def submit(self):
        # Add the form data to the Excel worksheet
        print("I'm in submit")
        name = self.name_entry.get()
        grade = self.grade_entry.get()
        gender = self.gender_var.get()
        school = self.school_var.get()
        class_ = self.class_var.get()
        new_values = [name, grade, gender, school, class_]

        #Do not submit if the name, grade, gender, school and class are not selected or entered.
        if name == "" or grade == "" or gender == "" or school == "" or class_ == "":
            messagebox.showerror("Error", "Please fill all the fields.")
            return  # Exit the submit method

        # Check if Excel file is open by another user
        while True:
            try:
                # open data.xlsx in the same directory as the script
                self.wb = openpyxl.load_workbook("data.xlsx")
                self.ws = self.wb.active

                if self.re_edit == True:
                    # replace the corresponding row from the Excel file
                    print("I'm in re_edit")
                    # print(self.inject_display_students.position)
                    self.replace_row(self.inject_display_students.position, new_values)
                    self.re_edit = False
                    break
                else:
                    # append the data to the Excel file
                    self.ws.append([name, grade, gender, school, class_])
                    # Save the workbook
                    self.wb.save("data.xlsx")
                    self.submit_button.config(state='normal')
                    break
            except PermissionError:
                self.submit_button.config(state='disabled')
                messagebox.showerror("Excel file is open", "Please close the Excel file and try again.")
                time.sleep(0.1)
                            # display the data in the GUI
        self.inject_display_students.refresh_student_display() # Call method to update display

        # Clear the form fields
        self.name_entry.delete(0, 'end')
        self.grade_entry.delete(0, 'end')
        self.gender_var.set(self.gender_options[0])
        self.school_var.set(self.school_options[0])
        self.class_var.set("")
 
    def clear(self):
        # Clear the form fields
        self.name_entry.delete(0, 'end')
        self.grade_entry.delete(0, 'end')
        self.gender_var.set(self.gender_options[0])
        self.school_var.set(self.school_options[0])
        self.class_var.set("")
        self.submit_button.config(state='normal')
        self.re_edit = False

    def delete(self):
        # Delete the selected row from the Excel file
        print("I'm in delete")
        if not self.re_edit:
            messagebox.showerror("Error", "Please select a row to delete.")
            return
        
        # Get the selected row

        position = self.inject_display_students.position 
        print("selected item", position)
        #position = int(position)+1
        # Delete the row
        print("DELETE ROW", position) 
        self.ws.delete_rows(position)
        # Save the workbook
        self.wb.save("data.xlsx")
        # Refresh the data viewer
        self.inject_display_students.refresh_student_display()  
        # Clear the form fields
        self.clear()

def main():
    root = tk.Tk()
    # Local main for testing Object1
    student_entry = StudentEntry(master=root)
    # Test Object1 functionality
    student_entry.mainloop()

if __name__ == "__main__":  
    main()