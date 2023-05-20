import tkinter as tk
from display_students import Display_students
from student_entry import StudentEntry
import openpyxl
from tkinter import messagebox

'''
This program:
Creates a GUI to enter student data into an Excel file.
Uses the openpyxl module to read and write Excel files.
Uses the tkinter module to create the GUI.
Creates a new data.xlsx file with headers if it does not exist.
Must be run from the same directory as the data.xlsx file.
Checks if the data.xlsx file is open by another user.
Uses school options taken from the schools.xlsx that must be in the same directory as the script with a sheet named "Schools".
    - Its columns of schools.xlsx must have a column named "School Name" and "Class 1", "Class 2", "Class 3", etc.
    - The schools.xlsx file must have a row for each school and all the classes.
It has nested dropdown menus for the schools and classes.
Shows the newly added data in the data viewer.
Allows the user to double-click a row in the viewer and the row data to go for editing in the entry form
The program is not ready to check if:
    - school and class are selected and valid.
    - name and grade are entered and valid.
 -The program as it is does not provide to:
- delete a row from the Excel file.
- sort the data in the viewer by any column.
- filter the data in the viewer by any column.
- search the data in the viewer by any column.

'''

    
if __name__ == "__main__":
    root = tk.Tk()

    # Check if data.xlsx exists.
    try:
        wb = openpyxl.load_workbook("data.xlsx")
    except:
        # Create the data.xlsx file.
        wb = openpyxl.Workbook()
        # Create the column headers.
        ws = wb.active
        # append the headers to the first row Name, Grade, Gender, OldSchool, OldSchoolClass
        ws.append(['Name','Gender','Grade','OldSchool','OldSchoolClass'])
        wb.save("data.xlsx")
        #

    # Check if the data.xlsx file is open by another user.
    try:
        wb = openpyxl.load_workbook("data.xlsx")
    except:
        messagebox.showerror("Error", "The data.xlsx file is open by another user.")
        exit()

    display_students = Display_students(master=root)

    student_entry = StudentEntry(master=root)
 
    student_entry.inject_display_students = display_students
    display_students.inject_student_entry = student_entry
    

    

    student_entry.mainloop()
    display_students.mainloop()