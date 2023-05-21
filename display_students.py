import openpyxl
import tkinter as tk
import tkinter.ttk as ttk

'''
This object creates a GUI to show the students in a scrollable viewer. 
The user can double-click on a row with a student data and these will be put in the entry form of the StudentEntry form, 
write the position of the row in the Excel file and change the re_edit status of StudentEntry object to True.   
Whenever a change to data.xlsx is made, the viewer is refreshed.
'''

class Display_students(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        #self.master.title("Excel Data Viewer")
        self.tree = ttk.Treeview(self.master)
        self.grid()


        # Create a canvas widget with a scrollbar
        self.canvas = tk.Canvas(self, width=500, height=300)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.scrollbar = tk.Scrollbar(self, orient=tk.VERTICAL, command=self.canvas.yview)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.config(yscrollcommand=self.scrollbar.set)
        self.canvas.bind('<Configure>', lambda e: self.canvas.config(scrollregion=self.canvas.bbox("all")))

        # Create a frame inside the canvas to hold the data
        self.data_frame = tk.Frame(self.canvas)
        self.canvas.create_window((0,0), window=self.data_frame, anchor=tk.NW)

        self.wb = openpyxl.load_workbook("data.xlsx")
        self.ws = self.wb.active
       
        # Create the treeview to display the data
        self.treeview = ttk.Treeview(self.data_frame, selectmode='browse')
        self.treeview.grid(row=1, column=0, padx=5, pady=5)

        # Define the columns
        self.treeview['columns'] = ('name', 'grade', 'gender', 'school', 'class')
        self.treeview.heading('#0', text='ID')
        self.treeview.column('#0', width=0, stretch='no')
        self.treeview.heading('name', text='Name')
        self.treeview.column('name', width=150, anchor='w')
        self.treeview.heading('grade', text='Grade')
        self.treeview.column('grade', width=75, anchor='center')
        self.treeview.heading('gender', text='Gender')
        self.treeview.column('gender', width=75, anchor='center')
        self.treeview.heading('school', text='School')
        self.treeview.column('school', width=150, anchor='w')
        self.treeview.heading('class', text='Class')
        self.treeview.column('class', width=75, anchor='center')

        # Display the data rows
        row_num = 1
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            print(row_num, row)
            self.treeview.insert('', 'end', text=row_num, values=row)
            row_num += 1

        # Bind the double-click event to the on_row_select method
        self.treeview.bind('<Double-1>', self.on_row_select)


    def on_row_select(self, event):
        # Get the selected row
        print("Hi I'm in on_row_select of display_students.py")
        print(event)
        self.inject_student_entry.re_edit = True
        self.selected_item1 = self.treeview.selection()[0]
        print("selected_item1", self.selected_item1)
        values = self.treeview.item(self.selected_item1, 'values')
        text = self.treeview.item(self.selected_item1, 'text')
        print("text row number :", text)
        self.position= int(text)+1
      
        # Fill the entry form with the selected row data
        self.inject_student_entry.name_entry.delete(0, 'end')
        self.inject_student_entry.name_entry.insert(0, values[0])
        self.inject_student_entry.grade_entry.delete(0, 'end')
        self.inject_student_entry.grade_entry.insert(0, values[1])
        self.inject_student_entry.gender_var.set(values[2])
        self.inject_student_entry.school_var.set(values[3])
        self.inject_student_entry.class_var.set(values[4])


    def refresh_student_display(self):
        # Display the data rows
        print("Hi I'm in refresh_student_display of display_students.py")
        self.treeview.delete(*self.treeview.get_children())
        # opend data.xlsx
        self.wb = openpyxl.load_workbook("data.xlsx")
        self.ws = self.wb.active
        row_num = 1
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            print(row_num, row)
            self.treeview.insert('', 'end', text=row_num, values=row)
            row_num += 1
        self.wb.save("data.xlsx")


def main():
    root = tk.Tk()
    # Local main for testing Object1
    display_students = Display_students(master=root)
    # Test Object1 functionality
    display_students.mainloop()

if __name__ == "__main__":
    main()